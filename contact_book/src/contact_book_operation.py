'''Contact Book Project'''

from abc import ABC, abstractmethod

import csv

from openpyxl import load_workbook, Workbook


class DataBaseService(ABC):
    """
    Database Service abstract class.

    Args:
        ABC (_type_): _description_
    """
    @abstractmethod
    def save_contact(self, name: str, phone_number: int, email: int):
        """
        Saves contact.

        Args:
            name (str): Contact name.
            phone_number (int): Contact phone number.
            email (int): Contact email address.
        """

    @abstractmethod
    def view_contacts(self, name: str) -> list:
        """
        Views a contact(s)

        Args:
            name (str): Contact name to be viewed.

        Returns:
            list: List of contact(s) to be viewed.
        """

    @abstractmethod
    def delete_contact(self, name: str):
        """
        Deletes a contact.

        Args:
            name (str): Contact name to be deleted.
        """

    @abstractmethod
    def clear_contact_book(self):
        """
        Clears all contacts.

        Raises:
            SyntaxError: Excel file not in correct format.
            SyntaxError: CSV file not in correct format.
        """


class ExcelDatabaseService(DataBaseService):
    """
    Excel Database Service class describing Excel functionality methods.

    Args:
        DataBaseService (class): Database service abstract class.
    """

    def __init__(self, file_name: str) -> None:
        self.file_name = file_name
        self.excel_header = ['Name', 'Phone Number', 'Email']

        try:
            self.workbook = load_workbook(self.file_name)
            self.sheet = self.workbook.active
            header_list = self.sheet[1]
            header_list = [cell_object.value for cell_object in header_list]
            print(header_list)
            if header_list == self.excel_header:
                print('Excel database loaded successfully.')
                return
            raise SyntaxError('Excel file not in correct format.')
        except FileNotFoundError:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(self.excel_header)

    def save_contact(self, name: str, phone_number: int, email: int):
        self.sheet.append([name, phone_number, email])
        self.workbook.save(self.file_name)

    def view_contacts(self, name: str) -> list:
        """
        Views a contact.

        Args:
            name (str): Contact name to be viewed.

        Returns:
            list: List of contacts to be viewed.
        """
        contacts = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if name.lower() in row[0].lower():
                contacts.append(list(row))

        return sorted(contacts)

    def delete_contact(self, name: str):
        """
        Deletes a contact.

        Args:
            name (str): Contact name to be deleted.
        """
        for idx, row in enumerate(self.sheet.iter_rows(values_only=True), 1):
            if row[0] == name:
                self.sheet.delete_rows(idx, 1)
                self.workbook.save(self.file_name)

    def clear_contact_book(self):
        """
        Clears all contacts.
        """
        while self.sheet.max_row > 1:
            self.sheet.delete_rows(2)
        self.workbook.save(self.file_name)


class CSVDatabaseService(DataBaseService):
    """
    CSV Database Service class describing CSV functionality methods.

    Args:
        DataBaseService (class): Database service abstract class.
    """

    def __init__(self, file_path: str) -> None:
        self.file_path = file_path
        self.csv_header = ['Name', 'Phone Number', 'Email']

        try:
            with open(self.file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                for row in reader:
                    if row[0] == 'Name' and row[1] == 'Phone Number' and row[2] == 'Email':
                        print('CSV database loaded successfully.')
                        return
                    raise SyntaxError('CSV file not in correct format.')
        except FileNotFoundError:
            with open(self.file_path, 'w', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(self.csv_header)

    def save_contact(self, name: str, phone_number: int, email: int):
        with open(self.file_path, 'a', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow([name, phone_number, email])

    def view_contacts(self, name: str) -> list:
        """
        Views contact(s).

        Args:
            name (str): Contact name to be viewed.

        Returns:
            list: List of contact(s) to be viewed.
        """
        contacts = []
        with open(self.file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            next(reader)
            for row in reader:
                if name in row:
                    contacts.append(row)

        return sorted(contacts)

    def delete_contact(self, name: str):
        """
        Deletes a contact.

        Args:
            name (str): Contact name to be deleted.

        """
        with open(self.file_path, 'r', encoding='utf-8') as file:
            contact_reader = csv.reader(file)
            contacts = [
                contact for contact in contact_reader if contact[0] != name]

        with open(self.file_path, 'w', encoding='utf-8') as file:
            contact_writer = csv.writer(file)
            contact_writer.writerows(contacts)

    def clear_contact_book(self):
        """
        Clears all contacts.
        """
        with open(self.file_path, 'w', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(self.csv_header)


class ContactBook:
    """
    Contact Book class for storing contact information like name, phone number
    and email address of a person.
    """

    def __init__(self, database_service: DataBaseService) -> None:
        self.database_service = database_service

    def add_contacts(self, name: str, phone_number: int, email: str) -> None:
        """
        Adds a contact.

        Args:
            name (str): Contact name to be added or updated.
            number (int): Number to be added or updated.
            email (str): Email to be added or updated.
        """
        self.database_service.save_contact(name, phone_number, email)

    def view_contacts(self, name: str) -> list:
        """
        Views a contact.

        Args:
            name (str): Contact name to be viewed.
        """
        return self.database_service.view_contacts(name)

    def delete_contact(self, name: str):
        """
        Deletes a contact.

        Args:
            name (str): Name to be deleted.
        """
        self.database_service.delete_contact(name)

    def clear_contact_book(self):
        """
        Clears all contacts.
        """
        self.database_service.clear_contact_book()

    def change_service_provider(self, new_service_provider: DataBaseService):
        """
        Changes database service provider.

        Args:
            new_service_provider (DataBaseService): New database service that
            has been selected.
        """
        self.database_service = new_service_provider
