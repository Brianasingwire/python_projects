'''Saving and Loading contacts to/from Excel Spreadsheet'''

import openpyxl


# def save_contacts(contacts, file_name):
#     '''Save contacts'''

#     book = Workbook()
#     book.create_sheet('Contacts')
#     sheet = book['Contacts']
#     print(book.sheetnames)

#     # sheet['A1'].value = 'Name'.upper()
#     # sheet['B1'].value = 'Phone Number'.upper()

#     for name, number in contacts.items():
#         sheet.append([name, number])

#     book.save(file_name)
#     print('Contacts saved successfully.')


def load_contacts(file_name='Contacts.xlsx'):
    """_summary_

    Args:
        file_name (str, optional): _description_. Defaults to 'Contacts.xlsx'.
    """

    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Phone Number', 'Email'])
    return (workbook, sheet)
