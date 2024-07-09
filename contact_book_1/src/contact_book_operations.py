'''Contact Book Project'''

from openpyxl import Workbook, load_workbook


class ContactBook:
    """
    Contact Book class describing a contact book.

    Attributes:
        file_name(str): File to which the contacts are to be saved and
                        loaded from.
    """

    def __init__(self, file_name='Contacts.xlsx') -> None:
        self.file_name = file_name

        try:
            self.workbook = load_workbook(self.file_name)
            self.sheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(['Name', 'Phone Number', 'Email'])

    def add_contact(self, name, phone, email):
        """
        Adding and updating contact to/in the Contact Book.

        Args:
            name (str): Name added/updated to/in Contact Book.
            phone (_type_): Phone number added/updated to/in Contact Book.
            email (str): Email address added/updated to/in Contact Book.
        """
        for idx, row in enumerate(self.sheet.iter_rows(values_only=True), start=1):
            if row[0] == name:
                self.sheet.cell(row=idx, column=2, value=phone)
                self.sheet.cell(row=idx, column=3, value=email)
                break
        else:
            self.sheet.append([name, phone, email])

    def view_contact(self, name):
        """
        Viewing a contact in the Contact Book.

        Args:
            name (str): Name of contact to be viewed.
        """
        for row in self.sheet.iter_rows(values_only=True):
            if row[0] == name:
                return {'Name': row[0], 'Phone': row[1], 'Email': row[2]}
        return None

    def remove_contact(self, name):
        """
        Removing a contact from the Contact Book.

        Args:
            name (str): Name of contact to be removed from Contact Book.
        """
        for idx, row in enumerate(self.sheet.iter_rows(values_only=True), start=1):
            if row[0] == name:
                self.sheet.delete_rows(idx, 1)
                break

    def save_contacts(self):
        """
        Saving contact(s) to the Contact Book.
        """
        self.workbook.save(self.file_name)

    def display_contacts(self):
        """
        Displaying contact(s) in the Contact Book.
        """
        for row in self.sheet.iter_rows(values_only=True):
            print(row)

    def contact_menu(self):
        """
        Printing out Contact Book menu.
        """

        print('')
        print('==========MENU==========')
        print('Press 1 to add/update a contact.')
        print('Press 2 to view a contact.')
        print('Press 3 to delete a contact.')
        print('Press 4 to view all contacts.')
        print('Press 5 to save contact book.')
        print('')
